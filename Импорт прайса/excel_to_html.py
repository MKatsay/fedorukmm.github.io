import os
import xlrd
import openpyxl
from html import escape

def excel_to_html(file_path):
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xls":
        return xls_to_html(file_path)
    elif ext == ".xlsx":
        return xlsx_to_html(file_path)
    else:
        raise ValueError("Формат не поддерживается (.xls или .xlsx)")


# ============================================================
#     ПАРСЕР XLS  (старые файлы Excel)
# ============================================================
def xls_to_html(path):
    book = xlrd.open_workbook(path, formatting_info=True)
    html = "<div class='excel-wrapper'>\n"

    for sheet in book.sheets():
        html += f"<h2>{escape(sheet.name)}</h2>\n"
        html += "<table border='1' cellspacing='0' cellpadding='4'>\n"

        # --- merged cells ---
        merges = sheet.merged_cells
        merge_map = {}
        for (r1, r2, c1, c2) in merges:
            merge_map[(r1, c1)] = (r2 - r1, c2 - c1)

        for r in range(sheet.nrows):
            html += "<tr>"
            for c in range(sheet.ncols):

                # пропускаем внутренние ячейки объединений
                if (r, c) not in merge_map:
                    if any(
                        r1 <= r < r1 + rs and c1 <= c < c1 + cs
                        for (r1, c1), (rs, cs) in merge_map.items()
                    ):
                        continue
                    rowspan = 1
                    colspan = 1
                else:
                    rowspan, colspan = merge_map[(r, c)]

                cell = sheet.cell(r, c)
                val = escape(str(cell.value).strip())

                # ---- СТИЛИ ----
                style = ""
                xf = book.xf_list[cell.xf_index]
                font = book.font_list[xf.font_index]

                if font.bold:
                    style += "font-weight:bold;"
                if font.italic:
                    style += "font-style:italic;"

                align = xf.alignment.hor_align
                if align == 1:
                    style += "text-align:center;"
                elif align == 3:
                    style += "text-align:right;"
                else:
                    style += "text-align:left;"

                html += f"<td rowspan='{rowspan}' colspan='{colspan}' style='{style}'>{val}</td>"

            html += "</tr>\n"

        html += "</table><br>\n"

    html += "</div>"
    return html


# ============================================================
#     ПАРСЕР XLSX (новые Excel)
# ============================================================
def xlsx_to_html(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    html = "<div class='excel-wrapper'>\n"

    for sheetname in wb.sheetnames:
        sh = wb[sheetname]

        html += f"<h2>{escape(sheetname)}</h2>\n"
        html += "<table border='1' cellspacing='0' cellpadding='4'>\n"

        # merged
        merges = sh.merged_cells.ranges
        merge_map = {}
        for m in merges:
            r1 = m.min_row - 1
            c1 = m.min_col - 1
            rs = m.max_row - m.min_row + 1
            cs = m.max_col - m.min_col + 1
            merge_map[(r1, c1)] = (rs, cs)

        max_r = sh.max_row
        max_c = sh.max_column

        for r in range(1, max_r + 1):
            html += "<tr>"
            for c in range(1, max_c + 1):
                key = (r - 1, c - 1)

                if key not in merge_map:
                    if any(
                        r1 <= (r-1) < r1 + rs and c1 <= (c-1) < c1 + cs
                        for (r1, c1), (rs, cs) in merge_map.items()
                    ):
                        continue
                    rowspan = 1
                    colspan = 1
                else:
                    rowspan, colspan = merge_map[key]

                cell = sh.cell(r, c)
                val = escape(str(cell.value if cell.value is not None else ""))

                style = ""
                if cell.font.bold:
                    style += "font-weight:bold;"
                if cell.font.italic:
                    style += "font-style:italic;"
                if cell.alignment.horizontal:
                    style += f"text-align:{cell.alignment.horizontal};"

                html += f"<td rowspan='{rowspan}' colspan='{colspan}' style='{style}'>{val}</td>"

            html += "</tr>\n"

        html += "</table><br>\n"

    html += "</div>"
    return html
